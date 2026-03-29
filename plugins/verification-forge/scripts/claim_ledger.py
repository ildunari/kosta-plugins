#!/usr/bin/env python3
"""Claim Ledger -- track and verify quantitative claims through document editing.

Subcommands:
  init        Create a new .claims.json ledger for a document
  add         Add a claim to an existing ledger
  update      Update claim status or fields
  verify      Check all claims and report pending/flagged
  export-md   Export ledger as a markdown table
  import-xlsx Import claims from an Excel spreadsheet

Requires Python 3.9+. Stdlib only; optionally: openpyxl (for xlsx import),
python-docx (for docx source reading).
"""

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from docx import Document as DocxDocument
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------

SCHEMA_VERSION = "1.0"

VALID_CLAIM_TYPES = {"direct", "derived", "inferential", "editorial", "source-stated"}
VALID_STATUSES = {"verified", "flagged", "pending"}
VALID_TIERS = {"standard", "high-stakes"}


def _now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _empty_claim(claim_id: str) -> dict:
    """Return a blank claim dict with all schema fields."""
    return {
        "id": claim_id,
        "text": "",
        "raw_value": "",
        "display_value": "",
        "unit": "",
        "source_file": "",
        "source_location": "",
        "citations": [],
        "claim_type": "direct",
        "verification_status": "pending",
        "verified_at": None,
        "notes": "",
    }


def _empty_ledger(document: str, tier: str = "standard") -> dict:
    return {
        "version": SCHEMA_VERSION,
        "document": document,
        "tier": tier,
        "created": _now_iso(),
        "claims": [],
    }


def _load_ledger(path: Path) -> dict:
    if not path.exists():
        print(f"Error: ledger not found: {path}", file=sys.stderr)
        sys.exit(1)
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        print(f"Error: invalid JSON in {path}: {exc}", file=sys.stderr)
        sys.exit(1)


def _save_ledger(path: Path, ledger: dict) -> None:
    path.write_text(json.dumps(ledger, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def _find_claim(ledger: dict, claim_id: str) -> dict | None:
    for c in ledger.get("claims", []):
        if c["id"] == claim_id:
            return c
    return None


def _read_text(path: Path) -> str:
    """Read text from .md, .txt, or .docx."""
    if path.suffix.lower() == ".docx":
        if not HAS_DOCX:
            print(f"  [warn] python-docx not installed -- skipping {path.name}", file=sys.stderr)
            return ""
        doc = DocxDocument(str(path))
        return "\n".join(p.text for p in doc.paragraphs)
    return path.read_text(encoding="utf-8", errors="replace")


# ---------------------------------------------------------------------------
# Subcommands
# ---------------------------------------------------------------------------

def cmd_init(args):
    """Create a new .claims.json for a document."""
    doc_path = Path(args.document)
    if not doc_path.exists():
        print(f"Warning: document '{doc_path}' does not exist yet.", file=sys.stderr)

    tier = args.tier if args.tier in VALID_TIERS else "standard"
    ledger = _empty_ledger(str(doc_path), tier)

    out = Path(args.output) if args.output else doc_path.with_suffix(".claims.json")
    if out.exists() and not args.force:
        print(f"Error: ledger already exists: {out} (use --force to overwrite)", file=sys.stderr)
        sys.exit(1)

    _save_ledger(out, ledger)
    print(f"Ledger created: {out} (tier={tier})")


def cmd_add(args):
    """Add a claim to an existing ledger."""
    ledger_path = Path(args.ledger)
    ledger = _load_ledger(ledger_path)

    if _find_claim(ledger, args.claim_id):
        print(f"Error: claim '{args.claim_id}' already exists in {ledger_path}", file=sys.stderr)
        sys.exit(1)

    claim = _empty_claim(args.claim_id)
    claim["text"] = args.text or ""
    claim["raw_value"] = args.raw_value or ""
    claim["display_value"] = args.display_value or args.raw_value or ""
    claim["unit"] = args.unit or ""
    claim["source_file"] = args.source_file or ""
    claim["source_location"] = args.source_location or ""
    claim["claim_type"] = args.claim_type if args.claim_type in VALID_CLAIM_TYPES else "direct"
    claim["verification_status"] = "pending"
    claim["notes"] = args.notes or ""

    if args.citations:
        claim["citations"] = [c.strip() for c in args.citations.split(",")]

    # For derived claims, store formula and raw_inputs if provided
    if args.claim_type == "derived":
        if args.formula:
            claim["formula"] = args.formula
        if args.raw_inputs:
            try:
                claim["raw_inputs"] = json.loads(args.raw_inputs)
            except json.JSONDecodeError:
                print(f"Warning: --raw-inputs is not valid JSON, storing as string", file=sys.stderr)
                claim["raw_inputs"] = args.raw_inputs

    ledger["claims"].append(claim)
    _save_ledger(ledger_path, ledger)
    print(f"Added claim {args.claim_id} to {ledger_path}")


def cmd_update(args):
    """Update claim status or fields."""
    ledger_path = Path(args.ledger)
    ledger = _load_ledger(ledger_path)

    claim = _find_claim(ledger, args.claim_id)
    if not claim:
        print(f"Error: claim '{args.claim_id}' not found in {ledger_path}", file=sys.stderr)
        sys.exit(1)

    updated_fields = []
    if args.status:
        if args.status not in VALID_STATUSES:
            print(f"Error: invalid status '{args.status}'. Must be one of: {', '.join(VALID_STATUSES)}", file=sys.stderr)
            sys.exit(1)
        claim["verification_status"] = args.status
        updated_fields.append("verification_status")
        if args.status == "verified":
            claim["verified_at"] = _now_iso()
            updated_fields.append("verified_at")

    if args.notes is not None:
        claim["notes"] = args.notes
        updated_fields.append("notes")
    if args.text is not None:
        claim["text"] = args.text
        updated_fields.append("text")
    if args.raw_value is not None:
        claim["raw_value"] = args.raw_value
        updated_fields.append("raw_value")
    if args.display_value is not None:
        claim["display_value"] = args.display_value
        updated_fields.append("display_value")
    if args.claim_type is not None:
        if args.claim_type not in VALID_CLAIM_TYPES:
            print(f"Error: invalid claim type '{args.claim_type}'.", file=sys.stderr)
            sys.exit(1)
        claim["claim_type"] = args.claim_type
        updated_fields.append("claim_type")

    if not updated_fields:
        print("Warning: no fields to update.", file=sys.stderr)
        return

    _save_ledger(ledger_path, ledger)
    print(f"Updated claim {args.claim_id}: {', '.join(updated_fields)}")


def cmd_verify(args):
    """Check all claims and report pending/flagged."""
    ledger_path = Path(args.ledger)
    ledger = _load_ledger(ledger_path)
    claims = ledger.get("claims", [])

    if not claims:
        print("Ledger has no claims.")
        return

    # Optionally verify against a document
    doc_text = ""
    if args.document:
        doc_path = Path(args.document)
        if doc_path.exists():
            doc_text = _read_text(doc_path)

    verified = 0
    flagged = 0
    pending = 0

    for c in claims:
        status = c.get("verification_status", "pending")

        # If we have a document, do value-matching verification
        if doc_text:
            display = c.get("display_value", "")
            raw = c.get("raw_value", "")
            if display and display in doc_text:
                c["verification_status"] = "verified"
                c["verified_at"] = _now_iso()
                status = "verified"
            elif raw and raw in doc_text:
                c["verification_status"] = "flagged"
                c["notes"] = (c.get("notes", "") + " Raw value present but display value differs.").strip()
                status = "flagged"
            elif display or raw:
                c["verification_status"] = "flagged"
                c["notes"] = (c.get("notes", "") + " Value not found in document.").strip()
                status = "flagged"

        if status == "verified":
            verified += 1
        elif status == "flagged":
            flagged += 1
        else:
            pending += 1

    total = len(claims)
    _save_ledger(ledger_path, ledger)

    print(f"Verification summary for {ledger_path}:")
    print(f"  Total claims: {total}")
    print(f"  Verified:     {verified}")
    print(f"  Flagged:      {flagged}")
    print(f"  Pending:      {pending}")

    if flagged > 0:
        print("\nFlagged claims:")
        for c in claims:
            if c.get("verification_status") == "flagged":
                print(f"  {c['id']}: {c.get('text', '')[:60]}...")
                if c.get("notes"):
                    print(f"    Note: {c['notes']}")

    if pending > 0:
        print("\nPending claims:")
        for c in claims:
            if c.get("verification_status") == "pending":
                print(f"  {c['id']}: {c.get('text', '')[:60]}...")

    # Exit non-zero if any claims are not verified
    if flagged > 0 or pending > 0:
        sys.exit(1)


def cmd_export_md(args):
    """Export ledger as a markdown table."""
    ledger_path = Path(args.ledger)
    ledger = _load_ledger(ledger_path)
    claims = ledger.get("claims", [])

    lines = [
        f"# Claim Ledger: {ledger.get('document', 'unknown')}",
        "",
        f"**Tier:** {ledger.get('tier', 'standard')} | **Version:** {ledger.get('version', '?')} | **Created:** {ledger.get('created', '?')}",
        "",
        "| ID | Text | Raw Value | Display | Unit | Type | Status | Source | Notes |",
        "|----|------|-----------|---------|------|------|--------|--------|-------|",
    ]

    for c in claims:
        text = c.get("text", "").replace("|", "\\|")[:80]
        lines.append(
            f"| {c['id']} "
            f"| {text} "
            f"| `{c.get('raw_value', '')}` "
            f"| {c.get('display_value', '')} "
            f"| {c.get('unit', '')} "
            f"| {c.get('claim_type', '')} "
            f"| **{c.get('verification_status', '')}** "
            f"| {c.get('source_file', '')} "
            f"| {c.get('notes', '')} |"
        )

    lines.append("")

    # Summary
    by_status = {}
    for c in claims:
        s = c.get("verification_status", "pending")
        by_status[s] = by_status.get(s, 0) + 1

    lines.append(f"**Total:** {len(claims)} claims")
    for s in ("verified", "flagged", "pending"):
        if by_status.get(s, 0) > 0:
            lines.append(f"- {s}: {by_status[s]}")
    lines.append("")

    output_text = "\n".join(lines)

    if args.output:
        out = Path(args.output)
        out.write_text(output_text, encoding="utf-8")
        print(f"Exported {len(claims)} claims to {out}")
    else:
        print(output_text)


def cmd_import_xlsx(args):
    """Import claims from an Excel spreadsheet into a ledger."""
    if not HAS_OPENPYXL:
        print("Error: openpyxl is required for xlsx import. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    xlsx_path = Path(args.spreadsheet)
    if not xlsx_path.exists():
        print(f"Error: spreadsheet not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    ledger_path = Path(args.ledger)
    ledger = _load_ledger(ledger_path)

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.active

    # Read header row to find column mappings
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value or "").strip().lower())

    # Map expected column names to indices
    col_map = {}
    expected = {
        "id": ["id", "claim_id", "claim id"],
        "text": ["text", "claim", "description", "claim text"],
        "raw_value": ["raw_value", "raw value", "value", "raw"],
        "display_value": ["display_value", "display value", "display"],
        "unit": ["unit", "units"],
        "source_file": ["source_file", "source file", "source", "file"],
        "source_location": ["source_location", "source location", "location"],
        "claim_type": ["claim_type", "claim type", "type"],
        "notes": ["notes", "note", "comment", "comments"],
    }

    for field, aliases in expected.items():
        for alias in aliases:
            if alias in headers:
                col_map[field] = headers.index(alias)
                break

    if "id" not in col_map:
        print("Error: spreadsheet must have an 'id' column.", file=sys.stderr)
        sys.exit(1)

    existing_ids = {c["id"] for c in ledger.get("claims", [])}
    added = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2):
        cells = [cell.value for cell in row]
        claim_id = str(cells[col_map["id"]] or "").strip()
        if not claim_id:
            continue
        if claim_id in existing_ids:
            skipped += 1
            continue

        claim = _empty_claim(claim_id)
        for field, idx in col_map.items():
            if field == "id":
                continue
            val = str(cells[idx] or "").strip() if idx < len(cells) and cells[idx] is not None else ""
            if field == "claim_type" and val not in VALID_CLAIM_TYPES:
                val = "direct"
            claim[field] = val

        claim["verification_status"] = "pending"
        ledger["claims"].append(claim)
        existing_ids.add(claim_id)
        added += 1

    wb.close()
    _save_ledger(ledger_path, ledger)
    print(f"Imported {added} claims from {xlsx_path} (skipped {skipped} duplicates)")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Claim Ledger -- track and verify quantitative claims through document editing.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""Examples:
  %(prog)s init report.md
  %(prog)s add report.claims.json --claim-id C-01 --text "72%% viability" --raw-value "72" --source-file data.xlsx
  %(prog)s update report.claims.json --claim-id C-01 --status verified
  %(prog)s verify report.claims.json --document report.md
  %(prog)s export-md report.claims.json --output claims_table.md
  %(prog)s import-xlsx data.xlsx report.claims.json
""",
    )
    sub = parser.add_subparsers(dest="command", required=True)

    # -- init --
    p_init = sub.add_parser("init", help="Create a new .claims.json ledger for a document")
    p_init.add_argument("document", help="Path to the document this ledger tracks")
    p_init.add_argument("--output", "-o", help="Output path (default: <document>.claims.json)")
    p_init.add_argument("--tier", choices=sorted(VALID_TIERS), default="standard", help="Verification tier")
    p_init.add_argument("--force", action="store_true", help="Overwrite existing ledger")
    p_init.set_defaults(func=cmd_init)

    # -- add --
    p_add = sub.add_parser("add", help="Add a claim to a ledger")
    p_add.add_argument("ledger", help="Path to .claims.json ledger")
    p_add.add_argument("--claim-id", required=True, help="Claim identifier (e.g. C-01)")
    p_add.add_argument("--text", help="Human-readable claim text")
    p_add.add_argument("--raw-value", help="Raw numeric value")
    p_add.add_argument("--display-value", help="Display-formatted value (default: same as raw-value)")
    p_add.add_argument("--unit", help="Unit of measurement")
    p_add.add_argument("--source-file", help="Source file the value comes from")
    p_add.add_argument("--source-location", help="Location within the source file")
    p_add.add_argument("--claim-type", choices=sorted(VALID_CLAIM_TYPES), default="direct", help="Type of claim")
    p_add.add_argument("--citations", help="Comma-separated citation IDs")
    p_add.add_argument("--notes", help="Free-text notes")
    p_add.add_argument("--formula", help="Formula for derived claims (e.g. 'a / b * 100')")
    p_add.add_argument("--raw-inputs", help="JSON dict of input variable names to values for derived claims")
    p_add.set_defaults(func=cmd_add)

    # -- update --
    p_upd = sub.add_parser("update", help="Update claim status or fields")
    p_upd.add_argument("ledger", help="Path to .claims.json ledger")
    p_upd.add_argument("--claim-id", required=True, help="Claim identifier to update")
    p_upd.add_argument("--status", choices=sorted(VALID_STATUSES), help="New verification status")
    p_upd.add_argument("--notes", help="Update notes")
    p_upd.add_argument("--text", help="Update claim text")
    p_upd.add_argument("--raw-value", help="Update raw value")
    p_upd.add_argument("--display-value", help="Update display value")
    p_upd.add_argument("--claim-type", choices=sorted(VALID_CLAIM_TYPES), help="Update claim type")
    p_upd.set_defaults(func=cmd_update)

    # -- verify --
    p_ver = sub.add_parser("verify", help="Check all claims and report pending/flagged")
    p_ver.add_argument("ledger", help="Path to .claims.json ledger")
    p_ver.add_argument("--document", help="Optional document to verify values against")
    p_ver.set_defaults(func=cmd_verify)

    # -- export-md --
    p_exp = sub.add_parser("export-md", help="Export ledger as a markdown table")
    p_exp.add_argument("ledger", help="Path to .claims.json ledger")
    p_exp.add_argument("--output", "-o", help="Output markdown file (default: stdout)")
    p_exp.set_defaults(func=cmd_export_md)

    # -- import-xlsx --
    p_imp = sub.add_parser("import-xlsx", help="Import claims from an Excel spreadsheet")
    p_imp.add_argument("spreadsheet", help="Path to .xlsx file")
    p_imp.add_argument("ledger", help="Path to .claims.json ledger to import into")
    p_imp.add_argument("--sheet", help="Sheet name (default: active sheet)")
    p_imp.set_defaults(func=cmd_import_xlsx)

    args = parser.parse_args()
    try:
        args.func(args)
    except KeyboardInterrupt:
        sys.exit(130)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
